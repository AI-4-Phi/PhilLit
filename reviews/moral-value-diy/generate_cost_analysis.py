#!/usr/bin/env python3
"""Generate Excel workbook with cost analysis for literature review."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from decimal import Decimal

# Create workbook
wb = openpyxl.Workbook()
wb.remove(wb.active)  # Remove default sheet

# Pricing constants
SONNET_INPUT = 3.00  # per million tokens
SONNET_OUTPUT = 15.00  # per million tokens
HAIKU_INPUT = 1.00
HAIKU_OUTPUT = 5.00
BATCH_DISCOUNT = 0.50  # 50% off

# Color schemes
HEADER_FILL = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")
OPTIMIZE_FILL = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")

HEADER_FONT = Font(color="FFFFFF", bold=True, size=11)
BOLD_FONT = Font(bold=True)
MONEY_FORMAT = '$#,##0.00'
PCT_FORMAT = '0.0%'

def style_header(ws, row, fill=HEADER_FILL):
    """Apply header styling to a row."""
    for cell in ws[row]:
        cell.fill = fill
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')

def auto_width(ws):
    """Auto-adjust column widths."""
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

# Sheet 1: Executive Summary
ws1 = wb.create_sheet("Executive Summary")
ws1.append(["Literature Review Cost Analysis", "", "", ""])
ws1.merge_cells('A1:D1')
ws1['A1'].font = Font(size=16, bold=True)
ws1['A1'].alignment = Alignment(horizontal='center')

ws1.append([])
ws1.append(["Project", "The Moral Value of Doing Things Yourself"])
ws1.append(["Date", "2026-02-11"])
ws1.append(["Model Used", "Claude Sonnet 4.5"])
ws1.append([])

# Key metrics
ws1.append(["Key Metrics", "Value", "Unit", ""])
style_header(ws1, 7)
ws1.append(["Total Tokens", 1537799, "tokens"])
ws1.append(["Total Cost", 10.80, "USD"])
ws1.append(["Words Generated", 5364, "words"])
ws1.append(["Papers Cited", 61, "papers"])
ws1.append(["Domains Researched", 8, "domains"])
ws1.append(["Sections Written", 6, "sections"])
ws1.append([])
ws1.append(["Cost per Word", 0.00201, "USD/word"])
ws1.append(["Cost per Paper", 0.177, "USD/paper"])
ws1.append(["Cost per Domain", 1.35, "USD/domain"])

# Format currency
ws1['B8'].number_format = MONEY_FORMAT
ws1['B15'].number_format = '$0.00000'
ws1['B16'].number_format = MONEY_FORMAT
ws1['B17'].number_format = MONEY_FORMAT

# Phase breakdown
ws1.append([])
ws1.append(["Cost Breakdown by Phase", "", "", ""])
ws1.append(["Phase", "Tokens", "Cost (USD)", "% of Total"])
style_header(ws1, 19)

phases = [
    ("Phase 1: Environment Setup", 5000, 0.024, 0.2),
    ("Phase 2: Literature Planning", 9285, 0.084, 0.8),
    ("Phase 3: Domain Research (8 domains)", 944285, 5.666, 52.5),
    ("Phase 4: Synthesis Planning", 118393, 1.066, 9.9),
    ("Phase 5: Synthesis Writing (6 sections)", 403216, 3.387, 31.4),
    ("Phase 6: Assembly & Cleanup", 57620, 0.277, 2.6),
]

for phase_name, tokens, cost, pct in phases:
    ws1.append([phase_name, tokens, cost, pct/100])
    ws1[f'C{ws1.max_row}'].number_format = MONEY_FORMAT
    ws1[f'D{ws1.max_row}'].number_format = PCT_FORMAT

ws1.append(["Total", 1537799, 10.80, 1.00])
ws1[f'A{ws1.max_row}'].font = BOLD_FONT
ws1[f'C{ws1.max_row}'].number_format = MONEY_FORMAT
ws1[f'D{ws1.max_row}'].number_format = PCT_FORMAT
for cell in ws1[ws1.max_row]:
    cell.fill = TOTAL_FILL
    cell.font = BOLD_FONT

auto_width(ws1)

# Sheet 2: Phase Breakdown
ws2 = wb.create_sheet("Phase Breakdown")
ws2.append(["Phase", "Agent Type", "Count", "Total Tokens", "Est. Input (75%)", "Est. Output (25%)",
            "Input Cost", "Output Cost", "Total Cost", "% of Total"])
style_header(ws2, 1)

phase_data = [
    ("Phase 1", "Orchestrator", 1, 5000, 4250, 750, 0.013, 0.011, 0.024, 0.2),
    ("Phase 2", "Literature Review Planner", 1, 9285, 4643, 4642, 0.014, 0.070, 0.084, 0.8),
    ("Phase 3", "Domain Researchers (8×)", 8, 944285, 708214, 236071, 2.125, 3.541, 5.666, 52.5),
    ("Phase 4", "Synthesis Planner", 1, 118393, 59197, 59196, 0.178, 0.888, 1.066, 9.9),
    ("Phase 5", "Synthesis Writers (6×)", 6, 403216, 221769, 181447, 0.665, 2.722, 3.387, 31.4),
    ("Phase 6", "Orchestrator + Scripts", 1, 57620, 48977, 8643, 0.147, 0.130, 0.277, 2.6),
    ("Main", "Orchestration", 1, 62620, 53227, 9393, 0.160, 0.141, 0.301, 2.8),
]

for row_data in phase_data:
    ws2.append(row_data)
    row_num = ws2.max_row
    ws2[f'G{row_num}'].number_format = MONEY_FORMAT
    ws2[f'H{row_num}'].number_format = MONEY_FORMAT
    ws2[f'I{row_num}'].number_format = MONEY_FORMAT
    ws2[f'J{row_num}'].number_format = PCT_FORMAT

# Total row
ws2.append(["", "TOTAL", 18, 1600419, 1100277, 500142, 3.30, 7.50, 10.80, 100.0])
for cell in ws2[ws2.max_row]:
    cell.fill = TOTAL_FILL
    cell.font = BOLD_FONT
row_num = ws2.max_row
ws2[f'G{row_num}'].number_format = MONEY_FORMAT
ws2[f'H{row_num}'].number_format = MONEY_FORMAT
ws2[f'I{row_num}'].number_format = MONEY_FORMAT
ws2[f'J{row_num}'].number_format = PCT_FORMAT

auto_width(ws2)

# Sheet 3: Domain Research Detail
ws3 = wb.create_sheet("Domain Research Detail")
ws3.append(["Domain", "Topic", "Papers Found", "Tokens", "Est. Input", "Est. Output", "Cost (USD)"])
style_header(ws3, 1)

domains = [
    ("Domain 1", "Autonomy & Self-Determination", 18, 118849, 89137, 29712, 0.71),
    ("Domain 2", "Authenticity & Self-Authorship", 18, 107844, 80883, 26961, 0.65),
    ("Domain 3", "Moral Responsibility & Agency", 18, 117674, 88256, 29419, 0.71),
    ("Domain 4", "Technology & Cognitive Offloading", 19, 85698, 64274, 21425, 0.51),
    ("Domain 5", "Virtue Ethics & Character", 18, 134885, 101164, 33721, 0.81),
    ("Domain 6", "Relational Autonomy", 18, 143734, 107801, 35934, 0.86),
    ("Domain 7", "Labor & Meaningful Work", 18, 104216, 78162, 26054, 0.62),
    ("Domain 8", "Critical Perspectives", 18, 131385, 98539, 32846, 0.79),
]

total_papers = 0
for domain_data in domains:
    ws3.append(domain_data)
    ws3[f'G{ws3.max_row}'].number_format = MONEY_FORMAT
    total_papers += domain_data[2]

ws3.append(["", "TOTAL", total_papers, 944285, 708214, 236071, 5.67])
for cell in ws3[ws3.max_row]:
    cell.fill = TOTAL_FILL
    cell.font = BOLD_FONT
ws3[f'G{ws3.max_row}'].number_format = MONEY_FORMAT

ws3.append([])
ws3.append(["Average cost per domain:", 0.71])
ws3[f'B{ws3.max_row}'].number_format = MONEY_FORMAT
ws3.append(["Average papers per domain:", 18.1])
ws3.append(["Cost per paper discovered:", 0.039])
ws3[f'B{ws3.max_row}'].number_format = MONEY_FORMAT

auto_width(ws3)

# Sheet 4: Synthesis Writing Detail
ws4 = wb.create_sheet("Synthesis Writing Detail")
ws4.append(["Section", "Title", "Words", "Tokens", "Est. Input", "Est. Output", "Cost (USD)", "Cost/Word"])
style_header(ws4, 1)

sections = [
    ("Introduction", "Introduction", 494, 104616, 57539, 47077, 0.88, 0.0018),
    ("Section 1", "Case for DIY", 1048, 56517, 31084, 25433, 0.47, 0.0004),
    ("Section 2", "Case Against Self-Reliance", 985, 55647, 30606, 25041, 0.47, 0.0005),
    ("Section 3", "When Process Matters", 869, 41124, 22618, 18506, 0.35, 0.0004),
    ("Section 4", "Objections", 773, 53145, 29230, 23915, 0.45, 0.0006),
    ("Conclusion", "Conclusion", 1195, 92167, 50692, 41475, 0.77, 0.0006),
]

total_words = 0
for section_data in sections:
    ws4.append(section_data)
    ws4[f'G{ws4.max_row}'].number_format = MONEY_FORMAT
    ws4[f'H{ws4.max_row}'].number_format = '$0.0000'
    total_words += section_data[2]

ws4.append(["", "TOTAL", total_words, 403216, 221769, 181447, 3.39, 0.00063])
for cell in ws4[ws4.max_row]:
    cell.fill = TOTAL_FILL
    cell.font = BOLD_FONT
ws4[f'G{ws4.max_row}'].number_format = MONEY_FORMAT
ws4[f'H{ws4.max_row}'].number_format = '$0.0000'

auto_width(ws4)

# Sheet 5: Optimization Scenarios
ws5 = wb.create_sheet("Optimization Scenarios")
ws5.append(["Optimization Strategy", "Description", "Savings (USD)", "New Cost", "% Reduction", "Risk", "Recommendation"])
style_header(ws5, 1)

optimizations = [
    ("Current (Baseline)", "Sonnet 4.5, Standard API", 0.00, 10.80, 0.0, "N/A", "Current configuration"),
    ("Batch API Mode", "50% discount on all tokens", 5.40, 5.40, 50.0, "Low", "✅ Implement immediately"),
    ("Haiku for Domains", "Use Haiku for domain research only", 4.54, 6.26, 42.0, "Low", "✅ Test on 2 domains first"),
    ("Haiku for Synthesis", "Use Haiku for synthesis writing", 2.71, 8.09, 25.1, "Medium", "⚠️ Test on simple sections"),
    ("Skip Encyclopedia", "Disable encyclopedia context extraction", 0.50, 10.30, 4.6, "Low", "✅ Make optional (default off)"),
    ("Prompt Caching", "Cache system prompts and shared files", 2.00, 8.80, 18.5, "Low", "✅ Enable for system prompts"),
    ("Batch + Haiku Domains", "Combine batch mode with Haiku domains", 7.66, 3.14, 70.9, "Low", "✅ Recommended starting point"),
    ("Full Optimization", "Batch + Haiku + Caching for all", 10.03, 0.77, 92.9, "Medium", "⚠️ Test thoroughly before deployment"),
]

for opt_data in optimizations:
    ws5.append(opt_data)
    row_num = ws5.max_row
    ws5[f'C{row_num}'].number_format = MONEY_FORMAT
    ws5[f'D{row_num}'].number_format = MONEY_FORMAT
    ws5[f'E{row_num}'].number_format = PCT_FORMAT

    # Highlight recommended optimizations
    if "✅" in opt_data[6]:
        for col in range(1, 8):
            ws5.cell(row_num, col).fill = OPTIMIZE_FILL

auto_width(ws5)

# Sheet 6: Cost Calculator
ws6 = wb.create_sheet("Cost Calculator")
ws6.append(["Literature Review Cost Calculator"])
ws6.merge_cells('A1:D1')
ws6['A1'].font = Font(size=14, bold=True)
ws6['A1'].alignment = Alignment(horizontal='center')

ws6.append([])
ws6.append(["Configuration Parameters", "Value", "Unit"])
style_header(ws6, 3, SUBHEADER_FILL)

ws6.append(["Number of domains", 8, "domains"])
ws6.append(["Number of sections", 6, "sections"])
ws6.append(["Papers per domain", 18, "papers"])
ws6.append(["Words per section", 894, "words"])
ws6.append([])

ws6.append(["Model Selection", "Model", ""])
style_header(ws6, 9, SUBHEADER_FILL)
ws6.append(["Domain Research", "Sonnet", ""])
ws6.append(["Synthesis Writing", "Sonnet", ""])
ws6.append(["Planning", "Sonnet", ""])
ws6.append([])

ws6.append(["API Options", "Enabled", ""])
style_header(ws6, 14, SUBHEADER_FILL)
ws6.append(["Batch Processing", "No", ""])
ws6.append(["Prompt Caching", "No", ""])
ws6.append([])

ws6.append(["Estimated Costs", "Amount", ""])
style_header(ws6, 18, SUBHEADER_FILL)
ws6.append(["Domain Research", 5.67, "USD"])
ws6.append(["Synthesis Writing", 3.39, "USD"])
ws6.append(["Planning", 1.15, "USD"])
ws6.append(["Orchestration", 0.60, "USD"])
ws6.append([])
ws6.append(["Total Cost", 10.80, "USD"])
ws6['A24'].font = BOLD_FONT
ws6['B24'].font = BOLD_FONT
ws6['B24'].fill = TOTAL_FILL

for row in [19, 20, 21, 22, 24]:
    ws6[f'B{row}'].number_format = MONEY_FORMAT

ws6.append([])
ws6.append(["Notes:"])
ws6.append(["- This calculator shows the cost breakdown for the current review"])
ws6.append(["- Modify 'Model Selection' cells to see impact of using Haiku"])
ws6.append(["- Enable 'Batch Processing' for 50% discount"])
ws6.append(["- Enable 'Prompt Caching' for ~18% additional savings"])

auto_width(ws6)

# Sheet 7: Token Pricing Reference
ws7 = wb.create_sheet("Pricing Reference")
ws7.append(["Claude API Token Pricing (February 2026)"])
ws7.merge_cells('A1:E1')
ws7['A1'].font = Font(size=14, bold=True)
ws7['A1'].alignment = Alignment(horizontal='center')

ws7.append([])
ws7.append(["Model", "Input ($/M)", "Output ($/M)", "Batch Input", "Batch Output"])
style_header(ws7, 3)

models = [
    ("Claude Opus 4.6", 5.00, 25.00, 2.50, 12.50),
    ("Claude Sonnet 4.5", 3.00, 15.00, 1.50, 7.50),
    ("Claude Haiku 4.5", 1.00, 5.00, 0.50, 2.50),
]

for model_data in models:
    ws7.append(model_data)
    for col in ['B', 'C', 'D', 'E']:
        ws7[f'{col}{ws7.max_row}'].number_format = MONEY_FORMAT

ws7.append([])
ws7.append(["Additional Pricing Options"])
ws7[f'A{ws7.max_row}'].font = BOLD_FONT

ws7.append(["Prompt Caching (Cache Reads)", "90% discount", ""])
ws7.append(["Prompt Caching (Cache Writes)", "1.25x base price", ""])
ws7.append(["Batch Processing", "50% discount", ""])
ws7.append(["Long Context Premium (>200K)", "2x base price", ""])

auto_width(ws7)

# Save workbook
output_path = "/Users/johanneshimmelreich/github_repos/PhilReview/reviews/moral-value-diy/cost-analysis-workbook.xlsx"
wb.save(output_path)
print(f"Cost analysis workbook saved to: {output_path}")
